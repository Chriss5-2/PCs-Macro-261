from __future__ import annotations

import argparse
import csv
import re
import sys
import unicodedata
from collections import Counter
from datetime import date, datetime, time
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - runtime dependency hint
    raise SystemExit(
        "Falta la dependencia 'openpyxl'. Instala con: pip install openpyxl"
    ) from exc


DEFAULT_INPUT = Path(__file__).resolve().parent / "CSV_base" / "xlsx_file.xlsx"
DEFAULT_OUTPUT = Path(__file__).resolve().parent / "CSV_final" / "csv_limpio.csv"


CONTROL_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")


def normalize_text(value: str) -> str:
    text = unicodedata.normalize("NFKC", value)
    text = text.replace("\ufeff", "")
    text = text.replace("\u200b", "")
    text = text.replace("\u200c", "")
    text = text.replace("\u200d", "")
    text = text.replace("\xa0", " ")
    text = CONTROL_CHARS.sub("", text)
    return text


def cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        return str(value)
    return normalize_text(str(value))


def detect_suspicious_text(rows: list[list[str]]) -> Counter[str]:
    findings: Counter[str] = Counter()
    for row in rows:
        for value in row:
            if "�" in value:
                findings["replacement_char"] += value.count("�")
            if "Ã" in value or "Â" in value:
                findings["possible_mojibake"] += 1
    return findings


def load_sheet_rows(xlsx_path: Path, sheet_ref: str | int | None) -> tuple[str, list[list[str]]]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_ref is None:
        worksheet = workbook[workbook.sheetnames[0]]
    elif isinstance(sheet_ref, int):
        worksheet = workbook[workbook.sheetnames[sheet_ref]]
    else:
        worksheet = workbook[sheet_ref]

    rows: list[list[str]] = []
    for raw_row in worksheet.iter_rows(values_only=True):
        rows.append([cell_to_text(cell) for cell in raw_row])

    return worksheet.title, rows


def write_csv(csv_path: Path, rows: list[list[str]], delimiter: str) -> None:
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle, delimiter=delimiter, quoting=csv.QUOTE_MINIMAL)
        writer.writerows(rows)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Lee un XLSX, limpia caracteres invisibles o problemáticos y genera un CSV en UTF-8-SIG."
    )
    parser.add_argument(
        "input_xlsx",
        nargs="?",
        type=Path,
        default=DEFAULT_INPUT,
        help=f"Ruta del XLSX de entrada (por defecto: {DEFAULT_INPUT})",
    )
    parser.add_argument(
        "output_csv",
        nargs="?",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Ruta del CSV de salida (por defecto: {DEFAULT_OUTPUT})",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Nombre de la hoja a exportar. Si no se indica, usa la primera.",
    )
    parser.add_argument(
        "--sheet-index",
        type=int,
        default=None,
        help="Índice de la hoja a exportar, basado en 0.",
    )
    parser.add_argument(
        "--delimiter",
        default=",",
        help="Separador del CSV de salida. Por defecto usa coma.",
    )
    args = parser.parse_args()

    if not args.input_xlsx.exists():
        print(f"No existe el archivo de entrada: {args.input_xlsx}", file=sys.stderr)
        return 1

    if len(args.delimiter) != 1:
        print("El delimitador debe ser un solo caracter.", file=sys.stderr)
        return 1

    sheet_ref: str | int | None
    if args.sheet_index is not None:
        sheet_ref = args.sheet_index
    else:
        sheet_ref = args.sheet

    sheet_name, rows = load_sheet_rows(args.input_xlsx, sheet_ref)
    write_csv(args.output_csv, rows, args.delimiter)

    suspicious = detect_suspicious_text(rows)
    print(f"Hoja exportada: {sheet_name}")
    print(f"Filas escritas: {len(rows)}")
    print(f"Columnas detectadas en la primera fila: {len(rows[0]) if rows else 0}")
    print(f"CSV generado: {args.output_csv}")
    if suspicious:
        print("Advertencia: todavía se detectaron patrones sospechosos en el texto exportado:")
        for key, count in suspicious.items():
            print(f"- {key}: {count}")
    else:
        print("Verificación básica OK: no se detectaron caracteres de reemplazo ni mojibake obvio.")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())